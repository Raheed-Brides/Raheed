from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import random
import string
from io import BytesIO
import phonenumbers
from phonenumbers import geocoder, carrier, timezone
import sqlite3

app = Flask(__name__, template_folder='.')
app.config['SECRET_KEY'] = 'your-secret-key-here-change-this-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///bookings.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Admin credentials
ADMIN_NAME = "omaradmin01119065057"
ADMIN_PHONE = "1119065057"

# Phone Validator Class
class PhoneValidator:
    """Utility class for phone number validation and formatting"""
    
    @staticmethod
    def validate_phone(phone_number, default_region='EG'):
        """
        Validate phone number and return formatted version
        
        Args:
            phone_number (str): Phone number to validate
            default_region (str): Default region code (EG for Egypt)
            
        Returns:
            dict: Contains is_valid, formatted_number, country, carrier info
        """
        try:
            # Parse the phone number
            parsed = phonenumbers.parse(phone_number, default_region)
            
            # Check if it's a valid number
            is_valid = phonenumbers.is_valid_number(parsed)
            
            if not is_valid:
                return {
                    'is_valid': False,
                    'error': 'Invalid phone number format',
                    'original': phone_number
                }
            
            # Get formatted number in international format
            formatted = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.INTERNATIONAL)
            
            # Get E164 format for database storage
            e164 = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
            
            # Get country information
            country = geocoder.country_name_for_number(parsed, "en")
            
            # Get carrier information
            carrier_name = carrier.name_for_number(parsed, "en")
            
            # Get timezone information
            timezones = timezone.time_zones_for_number(parsed)
            
            return {
                'is_valid': True,
                'formatted': formatted,
                'e164': e164,
                'country': country,
                'carrier': carrier_name,
                'timezones': timezones,
                'national': phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.NATIONAL)
            }
            
        except phonenumbers.NumberParseException as e:
            return {
                'is_valid': False,
                'error': str(e),
                'original': phone_number
            }
    
    @staticmethod
    def format_for_display(phone_number, default_region='EG'):
        """Format phone number for display purposes"""
        validation = PhoneValidator.validate_phone(phone_number, default_region)
        if validation['is_valid']:
            return validation['formatted']
        return phone_number
    
    @staticmethod
    def is_egyptian_mobile(phone_number):
        """Check if the number is a valid Egyptian mobile number"""
        try:
            parsed = phonenumbers.parse(phone_number, 'EG')
            return (phonenumbers.is_valid_number(parsed) and 
                    phonenumbers.number_type(parsed) == phonenumbers.PhoneNumberType.MOBILE)
        except:
            return False
    
    @staticmethod
    def extract_digits(phone_number):
        """Extract only digits from phone number"""
        return ''.join(filter(str.isdigit, phone_number))

# Booking Model
class Booking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    book_number = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='pending')

    def to_dict(self):
        return {
            'id': self.id,
            'book_number': self.book_number,
            'name': self.name,
            'phone': self.phone,
            'message': self.message,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'status': self.status
        }

# Utility Functions
def generate_book_number():
    """Generate unique booking number"""
    prefix = "RH"
    random_suffix = ''.join(random.choices(string.digits, k=6))
    return f"{prefix}-{random_suffix}"

def check_database_schema():
    """Check database schema and table structure"""
    conn = sqlite3.connect('instance/bookings.db')
    cursor = conn.cursor()
    
    # Get table schema
    cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='booking'")
    result = cursor.fetchone()
    
    schema_info = {}
    
    if result:
        schema_info['schema'] = result[0]
        
        # Get column information
        cursor.execute("PRAGMA table_info(booking)")
        columns = cursor.fetchall()
        schema_info['columns'] = columns
    else:
        schema_info['error'] = "booking table not found"
    
    conn.close()
    return schema_info

def check_duplicates():
    """Check for duplicate entries in the database"""
    conn = sqlite3.connect('instance/bookings.db')
    cursor = conn.cursor()
    
    # Check for duplicate book numbers
    cursor.execute("""
        SELECT book_number, COUNT(*) as count 
        FROM booking 
        GROUP BY book_number 
        HAVING COUNT(*) > 1
    """)
    duplicate_book_numbers = cursor.fetchall()
    
    # Check for duplicate phone numbers with same name
    cursor.execute("""
        SELECT phone, name, COUNT(*) as count 
        FROM booking 
        GROUP BY phone, name 
        HAVING COUNT(*) > 1
    """)
    duplicate_customers = cursor.fetchall()
    
    # Get all bookings for reference
    cursor.execute("SELECT id, book_number, name, phone, created_at FROM booking ORDER BY created_at DESC")
    all_bookings = cursor.fetchall()
    
    conn.close()
    
    return {
        'duplicate_book_numbers': duplicate_book_numbers,
        'duplicate_customers': duplicate_customers,
        'total_bookings': len(all_bookings),
        'all_bookings': all_bookings
    }

def test_phone_validation():
    """Test phone validation functionality"""
    test_cases = [
        # Egyptian numbers
        "+201119065057",
        "01119065057",
        "0111 906 5057",
        "+20 111 906 5057",
        
        # International numbers
        "+1234567890",
        "+44 20 7946 0958",
        "+1 (555) 123-4567",
        
        # Invalid numbers
        "123",
        "invalid",
        "+999999999999999",
    ]
    
    results = []
    
    for phone in test_cases:
        result = PhoneValidator.validate_phone(phone, default_region='EG')
        results.append({
            'phone': phone,
            'result': result
        })
    
    return results

# Routes
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/booking', methods=['POST'])
def create_booking():
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['name', 'phone']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return jsonify({
                'success': False,
                'message': f'Missing required fields: {", ".join(missing_fields)}',
                'field_error': True
            }), 400
        
        # Validate name (basic validation)
        name = data.get('name', '').strip()
        if len(name) < 2:
            return jsonify({
                'success': False,
                'message': 'Name must be at least 2 characters long',
                'field_error': True
            }), 400
        
        # Check if the credentials match admin credentials
        if data['name'] == ADMIN_NAME and data['phone'] == ADMIN_PHONE:
            # Redirect to customer page instead of booking
            return jsonify({
                'success': False,
                'redirect': True,
                'redirect_url': url_for('all_customers'),
                'message': 'Admin credentials detected. Redirecting to customer page...'
            }), 200
        
        # Validate phone number
        phone_validation = PhoneValidator.validate_phone(data['phone'], default_region='EG')
        if not phone_validation['is_valid']:
            return jsonify({
                'success': False,
                'message': phone_validation.get('error', 'Invalid phone number'),
                'phone_error': True
            }), 400
        
        # Format phone number to E164 for consistent storage
        formatted_phone = phone_validation['e164']
        
        # Generate unique book number
        book_number = generate_book_number()
        
        # Ensure unique book number
        while Booking.query.filter_by(book_number=book_number).first():
            book_number = generate_book_number()
        
        # Create new booking with validated and formatted phone
        booking = Booking(
            book_number=book_number,
            name=name,
            phone=formatted_phone,
            message=data.get('message', '').strip()[:500]  # Limit message length
        )
        
        db.session.add(booking)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Booking submitted successfully! We will contact you soon.',
            'book_number': book_number,
            'phone_formatted': formatted_phone
        }), 201
        
    except Exception as e:
        # Log the error for debugging
        app.logger.error(f"Booking submission error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'An error occurred while processing your booking. Please try again or contact us directly.',
            'error_details': str(e) if app.debug else None
        }), 500

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        
        if name == ADMIN_NAME and phone == ADMIN_PHONE:
            session['admin_logged_in'] = True
            return redirect(url_for('all_customers'))
        else:
            return render_template('admin_login.html', error='Invalid credentials')
    
    return render_template('admin_login.html')

@app.route('/admin/all-customers')
def all_customers():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    bookings = Booking.query.order_by(Booking.created_at.desc()).all()
    return render_template('all_customers.html', bookings=[b.to_dict() for b in bookings])

@app.route('/admin/booking/<int:booking_id>/status', methods=['PUT'])
def update_booking_status(booking_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    booking = Booking.query.get_or_404(booking_id)
    data = request.json
    booking.status = data.get('status', 'pending')
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/admin/export-customers-excel')
def export_customers_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        # Get all bookings
        bookings = Booking.query.order_by(Booking.created_at.desc()).all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Bookings"
        
        # Define headers
        headers = [
            'Book Number',
            'Customer Name',
            'Phone Number',
            'Message',
            'Booking Date',
            'Status'
        ]
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row, booking in enumerate(bookings, 2):
            ws.cell(row=row, column=1, value=booking.book_number)
            ws.cell(row=row, column=2, value=booking.name)
            ws.cell(row=row, column=3, value=booking.phone)
            ws.cell(row=row, column=4, value=booking.message or "")
            ws.cell(row=row, column=5, value=booking.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=6, value=booking.status.title())
        
        # Auto-adjust column widths with better handling
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = max(min(max_length + 2, 50), 10)  # Minimum 10, max 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for better readability
        ws.row_dimensions[1].height = 25
        for row_num in range(2, len(bookings) + 2):
            ws.row_dimensions[row_num].height = 20
        
        # Add borders and better formatting
        from openpyxl.styles import Border, Side
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells
        for row in ws.iter_rows(min_row=1, max_row=len(bookings)+1, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(vertical="center")
        
        # Create Excel file in memory
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        filename = f"customers_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/booking/<int:booking_id>', methods=['DELETE'])
def delete_booking(booking_id):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    booking = Booking.query.get_or_404(booking_id)
    try:
        db.session.delete(booking)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Customer deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

# Utility Routes for Testing and Maintenance
@app.route('/admin/utils/db-schema')
def admin_db_schema():
    """Admin route to check database schema"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    schema_info = check_database_schema()
    return jsonify(schema_info)

@app.route('/admin/utils/check-duplicates')
def admin_check_duplicates():
    """Admin route to check for duplicates"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    duplicates_info = check_duplicates()
    return jsonify(duplicates_info)

@app.route('/admin/utils/test-phone-validation')
def admin_test_phone_validation():
    """Admin route to test phone validation"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    test_results = test_phone_validation()
    return jsonify(test_results)

@app.route('/admin/utils')
def admin_utils():
    """Admin utilities dashboard"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    return render_template('admin_utils.html')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
